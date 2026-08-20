# syllabus/services/master_timetable_xlsx_builder.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font

from syllabus.services.timetable_pdf_builder import DAY_LABELS
from syllabus.services.timetable_xlsx_builder import _write_day_grid
from syllabus.services.master_timetable_pdf_builder import LABELS, _cell_text
from syllabus.models.master_timetable import ActivityType

_BOLD = Font(bold=True, size=14)


def build_master_timetable_xlsx(roster, language: str = "sw", teacher_id=None, class_level_id=None) -> bytes:
    labels = LABELS.get(language, LABELS["sw"])
    day_labels = DAY_LABELS.get(language, DAY_LABELS["sw"])

    period_slots = list(roster.period_slots.order_by("order"))

    slots_qs = roster.slots.select_related(
        "period_slot", "class_level", "activity_type",
        "assignment", "assignment__teacher", "assignment__subject_version__subject",
    )
    if teacher_id:
        slots_qs = slots_qs.filter(assignment__teacher_id=teacher_id)
    if class_level_id:
        slots_qs = slots_qs.filter(class_level_id=class_level_id)
    slots = list(slots_qs)

    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]

    ws.cell(row=1, column=1, value=f"{labels['title']} - {roster.name}").font = _BOLD
    row_cursor = 3

    routine_types = ActivityType.objects.filter(is_fixed_routine=True).order_by("default_order")
    if routine_types.exists():
        routine_text = ", ".join(
            (rt.label_sw if language == "sw" else rt.label_en) for rt in routine_types
        )
        ws.cell(row=row_cursor, column=1, value=f"{labels['routine_heading']}: {routine_text}")
        row_cursor += 2

    if not period_slots or not slots:
        ws.cell(row=row_cursor, column=1, value=labels["no_entries"])
    else:
        period_labels = [ps.label for ps in period_slots]
        period_times = {
            ps.label: (
                f"{ps.timestart.strftime('%H:%M')}-{ps.timefinish.strftime('%H:%M')}"
                if ps.timestart and ps.timefinish else ""
            )
            for ps in period_slots
        }

        by_day = {}
        whole_school_by_day = {}
        for s in slots:
            if s.class_level_id is None:
                whole_school_by_day.setdefault(s.day_of_week, {})[s.period_slot.label] = s
            else:
                by_day.setdefault(s.day_of_week, {}).setdefault(s.class_level.name, {})[s.period_slot.label] = s

        all_days = sorted(set(by_day.keys()) | set(whole_school_by_day.keys()))
        for day in all_days:
            day_whole_school = whole_school_by_day.get(day, {})
            class_rows = []
            for class_name, period_map in sorted(by_day.get(day, {}).items()):
                merged = {}
                for ps in period_slots:
                    if ps.is_break:
                        continue
                    override = day_whole_school.get(ps.label)
                    if override:
                        merged[ps.label] = _cell_text(override, language)
                    elif ps.label in period_map:
                        merged[ps.label] = _cell_text(period_map[ps.label], language)
                class_rows.append((class_name, merged))

            row_cursor = _write_day_grid(
                ws, row_cursor, day_labels.get(day, str(day)), period_labels, period_times,
                class_rows, labels["class_col"], labels["time_row"],
            )

    for col_idx in range(1, len(period_labels) + 2 if period_slots else 12):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
