# syllabus/services/timetable_xlsx_builder.py
"""
Excel (.xlsx) export ya ratiba - thamani tu (values), hakuna formula yoyote,
ili faili iwe salama kutumika/kuhaririwa popote bila hatari ya kuvunjika kwa
mahesabu ya Excel.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from syllabus.models.timetable import TimeTable
from syllabus.services.timetable_pdf_builder import DAY_LABELS, LABELS as TT_LABELS, _time_str
from syllabus.services.school_timetable_pdf_builder import _teacher_initials

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_day_grid(ws, row_cursor, day_label, periods, period_times, class_rows, class_col_label, time_row_label):
    ws.cell(row=row_cursor, column=1, value=day_label).font = _BOLD
    row_cursor += 1

    header = [class_col_label] + [str(p) for p in periods]
    for col, value in enumerate(header, start=1):
        cell = ws.cell(row=row_cursor, column=col, value=value)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER
    row_cursor += 1

    time_row = [time_row_label] + [period_times.get(p, "") for p in periods]
    for col, value in enumerate(time_row, start=1):
        cell = ws.cell(row=row_cursor, column=col, value=value)
        cell.alignment = _CENTER
        cell.border = _BORDER
    row_cursor += 1

    for class_name, period_map in class_rows:
        row = [class_name] + [period_map.get(p, "") for p in periods]
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=value)
            cell.border = _BORDER
        row_cursor += 1

    return row_cursor + 1


def build_teacher_timetable_xlsx(workstation, language: str = "sw") -> bytes:
    labels = TT_LABELS.get(language, TT_LABELS["sw"])
    day_labels = DAY_LABELS.get(language, DAY_LABELS["sw"])

    rows = list(
        TimeTable.objects.filter(
            workstation=workstation, day_of_week__isnull=False, period__isnull=False,
        )
        .select_related("subject_version__subject", "subject_version__class_level")
        .order_by("day_of_week", "period")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]

    ws.cell(row=1, column=1, value=f"{labels['title']} - {workstation.teacher.get_full_name() or workstation.teacher.username}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=workstation.school_name)
    row_cursor = 4

    if not rows:
        ws.cell(row=row_cursor, column=1, value=labels["no_entries"])
    else:
        periods = sorted({r.period for r in rows})
        period_times = {}
        for r in rows:
            if r.period not in period_times and r.timestart and r.timefinish:
                period_times[r.period] = f"{_time_str(r.timestart)}-{_time_str(r.timefinish)}"

        by_day = {}
        for r in rows:
            class_name = r.subject_version.class_level.name if r.subject_version.class_level else "-"
            by_day.setdefault(r.day_of_week, {}).setdefault(class_name, {})[r.period] = r.subject_version.subject.name

        for day in sorted(by_day.keys()):
            class_rows = sorted(by_day[day].items())
            row_cursor = _write_day_grid(
                ws, row_cursor, day_labels.get(day, str(day)), periods, period_times,
                class_rows, labels["class_col"], labels["time_row"],
            )

    for col_idx in range(1, 12):
        ws.column_dimensions[chr(64 + col_idx)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_school_timetable_xlsx(school_name: str, language: str = "sw") -> bytes:
    from syllabus.services.school_timetable_pdf_builder import LABELS as SCHOOL_LABELS

    labels = SCHOOL_LABELS.get(language, SCHOOL_LABELS["sw"])
    day_labels = DAY_LABELS.get(language, DAY_LABELS["sw"])

    rows = list(
        TimeTable.objects.filter(
            workstation__school_name=school_name, workstation__is_active=True,
            day_of_week__isnull=False, period__isnull=False,
        )
        .select_related("workstation__teacher", "subject_version__subject", "subject_version__class_level")
        .order_by("day_of_week", "period")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]

    ws.cell(row=1, column=1, value=f"{labels['title']} - {school_name}").font = Font(bold=True, size=14)
    row_cursor = 3

    if not rows:
        ws.cell(row=row_cursor, column=1, value=labels["no_entries"])
    else:
        periods = sorted({r.period for r in rows})
        period_times = {}
        for r in rows:
            if r.period not in period_times and r.timestart and r.timefinish:
                period_times[r.period] = f"{_time_str(r.timestart)}-{_time_str(r.timefinish)}"

        by_day = {}
        for r in rows:
            class_name = r.subject_version.class_level.name if r.subject_version.class_level else "-"
            subject = r.subject_version.subject.name
            initials = _teacher_initials(r.workstation.teacher)
            by_day.setdefault(r.day_of_week, {}).setdefault(class_name, {})[r.period] = f"{subject} ({initials})"

        for day in sorted(by_day.keys()):
            class_rows = sorted(by_day[day].items())
            row_cursor = _write_day_grid(
                ws, row_cursor, day_labels.get(day, str(day)), periods, period_times,
                class_rows, labels["class_col"], labels["time_row"],
            )

    for col_idx in range(1, 12):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
