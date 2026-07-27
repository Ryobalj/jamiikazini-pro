# syllabus/services/exam_results_xlsx_builder.py
"""
Excel (.xlsx) export ya matokeo ya mtihani - thamani tu (values), hakuna
formula yoyote - alama/daraja/wastani/nafasi vyote vimekokotolewa Python
kabla ya kuandikwa, hivyo ni salama kuhaririwa/kutumika tena bila hatari ya
Excel kuvunja/kupitisha upya mahesabu.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from syllabus.services.exam_results_service import compute_class_results
from syllabus.services.exam_results_pdf_builder import LABELS

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_title(ws, exam, labels, title):
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=exam.workstation.school_name)
    ws.cell(row=3, column=1, value=f"{exam.name} - {exam.class_level.name} - {exam.get_term_display()} {exam.year}")


def build_subject_result_xlsx(exam, subject, language: str = "sw") -> bytes:
    labels = LABELS.get(language, LABELS["sw"])
    results = compute_class_results(exam)

    wb = Workbook()
    ws = wb.active
    ws.title = subject.name[:31]

    _write_title(ws, exam, labels, f"{labels['title']} - {subject.name}")

    header_row = 5
    for col, value in enumerate([labels["na"], labels["student"], labels["score"], labels["grade"], labels["rank"]], start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER

    if not results["students"]:
        ws.cell(row=header_row + 1, column=1, value=labels["no_students"])
    else:
        rows = sorted(
            results["students"],
            key=lambda r: (
                r["per_subject"].get(subject.id, {}).get("rank") is None,
                r["per_subject"].get(subject.id, {}).get("rank") or 0,
            ),
        )
        for i, r in enumerate(rows, start=1):
            entry = r["per_subject"].get(subject.id, {})
            values = [
                i,
                r["student"].full_name,
                float(entry["score"]) if entry.get("score") is not None else None,
                entry.get("grade") or "-",
                entry.get("rank") if entry.get("rank") is not None else "-",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=header_row + i, column=col, value=value)
                cell.border = _BORDER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    for col in ("C", "D", "E"):
        ws.column_dimensions[col].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_class_report_xlsx(exam, language: str = "sw") -> bytes:
    labels = LABELS.get(language, LABELS["sw"])
    results = compute_class_results(exam)
    subjects = results["subjects"]

    wb = Workbook()
    ws = wb.active
    ws.title = exam.class_level.name[:31]

    _write_title(ws, exam, labels, f"{labels['title']} - {exam.class_level.name}")

    header_row_1 = 5
    header_row_2 = 6

    ws.cell(row=header_row_1, column=1, value=labels["na"]).font = _BOLD
    ws.cell(row=header_row_1, column=2, value=labels["student"]).font = _BOLD
    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)

    col = 3
    for subject in subjects:
        ws.cell(row=header_row_1, column=col, value=subject.name).font = _BOLD
        ws.merge_cells(start_row=header_row_1, start_column=col, end_row=header_row_1, end_column=col + 2)
        ws.cell(row=header_row_2, column=col, value=labels["score"]).font = _BOLD
        ws.cell(row=header_row_2, column=col + 1, value=labels["grade"]).font = _BOLD
        ws.cell(row=header_row_2, column=col + 2, value=labels["rank"]).font = _BOLD
        col += 3

    summary_labels = [labels["total"], labels["average"], labels["grade"], labels["rank"]]
    for offset, label in enumerate(summary_labels):
        ws.cell(row=header_row_1, column=col + offset, value=label).font = _BOLD
        ws.merge_cells(start_row=header_row_1, start_column=col + offset, end_row=header_row_2, end_column=col + offset)

    for c in range(1, col + len(summary_labels)):
        for r in (header_row_1, header_row_2):
            cell = ws.cell(row=r, column=c)
            cell.alignment = _CENTER
            cell.border = _BORDER

    if not results["students"]:
        ws.cell(row=header_row_2 + 1, column=1, value=labels["no_students"])
    else:
        for i, r in enumerate(results["students"], start=1):
            row_num = header_row_2 + i
            ws.cell(row=row_num, column=1, value=i).border = _BORDER
            ws.cell(row=row_num, column=2, value=r["student"].full_name).border = _BORDER

            col = 3
            for subject in subjects:
                entry = r["per_subject"].get(subject.id, {})
                ws.cell(row=row_num, column=col, value=float(entry["score"]) if entry.get("score") is not None else None).border = _BORDER
                ws.cell(row=row_num, column=col + 1, value=entry.get("grade") or "-").border = _BORDER
                ws.cell(row=row_num, column=col + 2, value=entry.get("rank") if entry.get("rank") is not None else "-").border = _BORDER
                col += 3

            ws.cell(row=row_num, column=col, value=float(r["total"]) if r["total"] is not None else None).border = _BORDER
            ws.cell(row=row_num, column=col + 1, value=round(float(r["average"]), 2) if r["average"] is not None else None).border = _BORDER
            ws.cell(row=row_num, column=col + 2, value=r["overall_grade"] or "-").border = _BORDER
            ws.cell(row=row_num, column=col + 3, value=r["overall_rank"] if r["overall_rank"] is not None else "-").border = _BORDER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    for col_idx in range(3, 3 + len(subjects) * 3 + 4):
        letter = chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)
        ws.column_dimensions[letter].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
